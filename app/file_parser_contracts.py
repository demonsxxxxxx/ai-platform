from __future__ import annotations

import codecs
import hashlib
import io
import json
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any, Literal
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from pydantic import BaseModel, ConfigDict, Field, field_validator

from app.context_manifest import CONTEXT_MANIFEST_SCHEMA_VERSION, utf8_token_estimate
from app.validation import assert_safe_id


ATTACHMENT_PREPROCESSING_SCHEMA_VERSION = "ai-platform.attachment-preprocessing.v1"
ATTACHMENT_CONTEXT_SCHEMA_VERSION = "ai-platform.attachment-context.v1"
XLSX_PARSER_ID = "ai-platform.xlsx.openpyxl"
XLSX_PARSER_VERSION = "1"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_XLSX_FILE_BYTES = 1024 * 1024
MAX_XLSX_SHEETS = 16
MAX_XLSX_ROWS_PER_SHEET = 100
MAX_XLSX_COLUMNS_PER_SHEET = 32
MAX_XLSX_CELLS = 2048
MAX_XLSX_CELL_CHARS = 256
MAX_XLSX_PROMPT_CHARS = 16_000
MAX_XLSX_PROMPT_TOKENS = 24_000
MAX_XLSX_ZIP_ENTRIES = 2000
MAX_XLSX_ZIP_ENTRY_BYTES = 8 * 1024 * 1024
MAX_XLSX_ZIP_TOTAL_BYTES = 32 * 1024 * 1024
MAX_XLSX_ZIP_COMPRESSION_RATIO = 100

_SPREADSHEET_XML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_WORKBOOK_SHEET_TAG = f"{{{_SPREADSHEET_XML_NAMESPACE}}}sheet"
_WORKBOOK_SHEET_RELATIONSHIP_ID = f"{{{_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE}}}id"
_WORKBOOK_ROOT_TAG = f"{{{_SPREADSHEET_XML_NAMESPACE}}}workbook"
_WORKBOOK_SHEETS_TAG = f"{{{_SPREADSHEET_XML_NAMESPACE}}}sheets"
_PACKAGE_RELATIONSHIPS_ROOT_TAG = f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationships"
_PACKAGE_RELATIONSHIP_TAG = f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"
_WORKSHEET_RELATIONSHIP_TYPE = f"{_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE}/worksheet"
_CONTENT_TYPES_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/content-types"
_CONTENT_TYPES_ROOT_TAG = f"{{{_CONTENT_TYPES_NAMESPACE}}}Types"
_CONTENT_TYPES_DEFAULT_TAG = f"{{{_CONTENT_TYPES_NAMESPACE}}}Default"
_CONTENT_TYPES_OVERRIDE_TAG = f"{{{_CONTENT_TYPES_NAMESPACE}}}Override"
_CANONICAL_WORKBOOK_PART = "xl/workbook.xml"
_XLSX_WORKBOOK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
_OPENPYXL_WORKBOOK_CONTENT_TYPES = frozenset(
    {
        "application/vnd.ms-excel.template.macroEnabled.main+xml",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
        _XLSX_WORKBOOK_CONTENT_TYPE,
    }
)
_FORBIDDEN_XML_DECLARATIONS = (b"<!DOCTYPE", b"<!ENTITY")
_FORBIDDEN_XML_DECLARATION_TEXT = tuple(token.decode("ascii") for token in _FORBIDDEN_XML_DECLARATIONS)

_SUPPORTED_XLSX_EXTENSIONS = frozenset({".xlsx"})
_UNSUPPORTED_WORKBOOK_EXTENSIONS = frozenset({".xls", ".xlsb", ".xlsm", ".ods"})
_UNSUPPORTED_WORKBOOK_CONTENT_TYPES = frozenset(
    {
        "application/vnd.ms-excel",
        "application/vnd.ms-excel.sheet.binary.macroenabled.12",
        "application/vnd.ms-excel.sheet.macroenabled.12",
        "application/vnd.oasis.opendocument.spreadsheet",
    }
)


class AttachmentPreprocessingError(ValueError):
    """Fail-closed attachment preprocessing error with a stable machine code."""

    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


class AttachmentParserRequirement(BaseModel):
    """Server-owned preprocessing requirement for one run attachment."""

    model_config = ConfigDict(extra="forbid")

    file_id: str
    file_name: str
    extension: str
    content_type: str
    parser_id: str
    parser_version: str
    supported: bool = True
    max_bytes: int = Field(ge=1)
    expected_byte_count: int | None = Field(default=None, ge=0)
    expected_sha256: str | None = None

    @field_validator("file_id")
    @classmethod
    def validate_file_id(cls, value: str):
        return assert_safe_id(value, "file_id")

    @field_validator("expected_sha256")
    @classmethod
    def validate_expected_sha256(cls, value: str | None):
        if value is None:
            return None
        normalized = str(value).lower()
        if len(normalized) != 64 or any(character not in "0123456789abcdef" for character in normalized):
            raise ValueError("expected_sha256 must be 64 lowercase hexadecimal characters")
        return normalized


class MaterializedAttachmentFact(BaseModel):
    """Ordered server fact captured from bytes fetched for one exact file ID."""

    model_config = ConfigDict(extra="forbid")

    file_id: str
    file_name: str
    content_type: str = ""
    byte_count: int = Field(ge=0)
    sha256: str

    @field_validator("file_id")
    @classmethod
    def validate_file_id(cls, value: str):
        return assert_safe_id(value, "file_id")

    @field_validator("sha256")
    @classmethod
    def validate_sha256(cls, value: str):
        normalized = str(value or "").lower()
        if len(normalized) != 64 or any(character not in "0123456789abcdef" for character in normalized):
            raise ValueError("sha256 must be 64 lowercase hexadecimal characters")
        return normalized


class AttachmentParserEvidence(BaseModel):
    """Bounded positive evidence emitted by a platform attachment parser."""

    model_config = ConfigDict(extra="forbid")

    file_id: str
    parser_id: str
    parser_version: str
    content_type: str
    extension: str
    byte_count: int = Field(ge=0)
    sha256: str
    sheet_count: int = Field(ge=0)
    sheets_processed: int = Field(ge=0)
    cells_examined: int = Field(ge=0)
    nonempty_cells: int = Field(ge=0)
    rows_emitted: int = Field(ge=0)
    truncated: bool
    status: Literal["parsed"]

    @field_validator("file_id")
    @classmethod
    def validate_file_id(cls, value: str):
        return assert_safe_id(value, "file_id")

    @field_validator("sha256")
    @classmethod
    def validate_sha256(cls, value: str):
        normalized = str(value or "").lower()
        if len(normalized) != 64 or any(character not in "0123456789abcdef" for character in normalized):
            raise ValueError("sha256 must be 64 lowercase hexadecimal characters")
        return normalized


class ParsedAttachmentContext(BaseModel):
    """Typed, bounded attachment content forwarded separately from user text."""

    model_config = ConfigDict(extra="forbid")

    evidence: AttachmentParserEvidence
    content: dict[str, Any]


@dataclass(frozen=True)
class AttachmentParserSpec:
    """Immutable platform-owned parser registration; never loaded from a Skill."""

    parser_id: str
    parser_version: str
    extensions: frozenset[str]
    content_types: frozenset[str]
    max_bytes: int


XLSX_PARSER_SPEC = AttachmentParserSpec(
    parser_id=XLSX_PARSER_ID,
    parser_version=XLSX_PARSER_VERSION,
    extensions=_SUPPORTED_XLSX_EXTENSIONS,
    content_types=frozenset({XLSX_CONTENT_TYPE}),
    max_bytes=MAX_XLSX_FILE_BYTES,
)
ATTACHMENT_PARSER_REGISTRY = (XLSX_PARSER_SPEC,)


def _normalized_extension(file_name: object) -> str:
    safe_name = PurePosixPath(str(file_name or "").replace("\\", "/")).name
    return Path(safe_name).suffix.casefold()


def _normalized_content_type(content_type: object) -> str:
    return str(content_type or "").split(";", 1)[0].strip().casefold()


def parser_spec_for_attachment(
    *,
    file_name: object,
    content_type: object = "",
) -> AttachmentParserSpec | None:
    """Resolve a parser only from the immutable platform registry."""

    extension = _normalized_extension(file_name)
    normalized_type = _normalized_content_type(content_type)
    for spec in ATTACHMENT_PARSER_REGISTRY:
        if extension in spec.extensions or normalized_type in spec.content_types:
            return spec
    return None


def is_known_binary_workbook(*, file_name: object, content_type: object = "") -> bool:
    """Return whether a file must be staged/parsed instead of text-decoded."""

    extension = _normalized_extension(file_name)
    normalized_type = _normalized_content_type(content_type)
    return bool(
        parser_spec_for_attachment(file_name=file_name, content_type=content_type)
        or extension in _UNSUPPORTED_WORKBOOK_EXTENSIONS
        or normalized_type in _UNSUPPORTED_WORKBOOK_CONTENT_TYPES
    )


def dispatched_context_file_ids(manifest: object) -> frozenset[str]:
    """Return the immutable exact file-ID authority dispatched to the sandbox."""

    if not isinstance(manifest, dict) or manifest.get("schema_version") != CONTEXT_MANIFEST_SCHEMA_VERSION:
        return frozenset()
    rows = manifest.get("files")
    if not isinstance(rows, list):
        return frozenset()
    return frozenset(
        str(row.get("file_id") or "").strip()
        for row in rows
        if isinstance(row, dict) and str(row.get("file_id") or "").strip()
    )


def build_attachment_preprocessing_contract(
    *,
    file_ids: list[str] | None = None,
    file_names: list[str] | None = None,
    content_types: list[str] | None = None,
    attachment_facts: list[MaterializedAttachmentFact | dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build server-owned parser requirements from ordered attachment facts."""

    requirements: list[AttachmentParserRequirement] = []
    normalized_facts: list[MaterializedAttachmentFact | None]
    if attachment_facts is not None:
        try:
            normalized_facts = [
                fact
                if isinstance(fact, MaterializedAttachmentFact)
                else MaterializedAttachmentFact.model_validate(fact)
                for fact in attachment_facts
            ]
        except Exception as exc:
            raise AttachmentPreprocessingError("attachment_materialized_fact_invalid") from exc
        ordered_file_ids = [fact.file_id for fact in normalized_facts if fact is not None]
        ordered_file_names = [fact.file_name for fact in normalized_facts if fact is not None]
        ordered_content_types = [fact.content_type for fact in normalized_facts if fact is not None]
    else:
        ordered_file_ids = list(file_ids or [])
        ordered_file_names = list(file_names or [])
        ordered_content_types = list(content_types or [])
        normalized_facts = [None for _name in ordered_file_names]
    for index, file_name in enumerate(ordered_file_names):
        extension = _normalized_extension(file_name)
        declared_content_type = (
            _normalized_content_type(ordered_content_types[index])
            if index < len(ordered_content_types)
            else ""
        )
        spec = parser_spec_for_attachment(
            file_name=file_name,
            content_type=declared_content_type,
        )
        if (
            spec is None
            and extension not in _UNSUPPORTED_WORKBOOK_EXTENSIONS
            and declared_content_type not in _UNSUPPORTED_WORKBOOK_CONTENT_TYPES
        ):
            continue
        if index >= len(ordered_file_ids):
            raise AttachmentPreprocessingError("attachment_parser_file_mapping_invalid")
        fact = normalized_facts[index]
        requirement = AttachmentParserRequirement(
            file_id=ordered_file_ids[index],
            file_name=PurePosixPath(str(file_name).replace("\\", "/")).name,
            extension=extension,
            content_type=(
                declared_content_type
                or (XLSX_CONTENT_TYPE if spec is not None else "application/octet-stream")
            ),
            parser_id=spec.parser_id if spec is not None else "unsupported",
            parser_version=spec.parser_version if spec is not None else "0",
            supported=spec is not None,
            max_bytes=spec.max_bytes if spec is not None else 1,
            expected_byte_count=fact.byte_count if fact is not None else None,
            expected_sha256=fact.sha256 if fact is not None else None,
        )
        requirements.append(requirement)
    return {
        "schema_version": ATTACHMENT_PREPROCESSING_SCHEMA_VERSION,
        "requirements": [requirement.model_dump(mode="json") for requirement in requirements],
    }


def attachment_requirements_from_contract(value: object) -> list[AttachmentParserRequirement]:
    """Validate that a runtime contract exactly matches the server registry."""

    if not isinstance(value, dict):
        return []
    if value.get("schema_version") != ATTACHMENT_PREPROCESSING_SCHEMA_VERSION:
        raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
    raw_requirements = value.get("requirements")
    if not isinstance(raw_requirements, list):
        raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
    requirements: list[AttachmentParserRequirement] = []
    seen_file_ids: set[str] = set()
    for raw in raw_requirements:
        try:
            requirement = AttachmentParserRequirement.model_validate(raw)
        except Exception as exc:
            raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid") from exc
        if requirement.file_id in seen_file_ids:
            raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
        seen_file_ids.add(requirement.file_id)
        rebuilt = build_attachment_preprocessing_contract(
            file_ids=[requirement.file_id],
            file_names=[requirement.file_name],
            content_types=[requirement.content_type],
        )["requirements"]
        if len(rebuilt) != 1:
            raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
        expected = rebuilt[0]
        actual = requirement.model_dump(mode="json")
        for key in (
            "file_id",
            "file_name",
            "extension",
            "content_type",
            "parser_id",
            "parser_version",
            "supported",
            "max_bytes",
        ):
            if actual[key] != expected[key]:
                raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
        if (requirement.expected_byte_count is None) != (requirement.expected_sha256 is None):
            raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
        requirements.append(requirement)
    return requirements


def _xml_multibyte_encoding(prefix: bytes) -> str | None:
    if prefix.startswith((codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE)):
        return "utf-32"
    if prefix.startswith((codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)):
        return "utf-16"
    if prefix.startswith(b"\x00\x00\x00<"):
        return "utf-32-be"
    if prefix.startswith(b"<\x00\x00\x00"):
        return "utf-32-le"
    if prefix.startswith(b"\x00<"):
        return "utf-16-be"
    if prefix.startswith(b"<\x00"):
        return "utf-16-le"
    return None


def _assert_xml_payload_has_no_dtd_or_entity(payload: bytes) -> None:
    try:
        encoding = _xml_multibyte_encoding(payload)
        if encoding is not None:
            text = payload.decode(encoding, errors="strict").upper()
            if any(token in text for token in _FORBIDDEN_XML_DECLARATION_TEXT):
                raise AttachmentPreprocessingError("xlsx_xml_entities_unsupported")
            return
        probe = payload.removeprefix(codecs.BOM_UTF8).lstrip(b" \t\r\n")
        if b"\x00" in payload or (probe and not probe.startswith(b"<")):
            raise AttachmentPreprocessingError("xlsx_xml_encoding_unsupported")
        upper_payload = payload.upper()
        if any(token in upper_payload for token in _FORBIDDEN_XML_DECLARATIONS):
            raise AttachmentPreprocessingError("xlsx_xml_entities_unsupported")
    except AttachmentPreprocessingError:
        raise
    except (UnicodeError, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc


def _check_classified_xml_entry(archive: ZipFile, entry: Any) -> bool:
    normalized_name = str(entry.filename).replace("\\", "/").casefold()
    if not normalized_name.endswith((".xml", ".rels")):
        return False
    try:
        payload = archive.read(entry)
    except (BadZipFile, OSError, RuntimeError, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    _assert_xml_payload_has_no_dtd_or_entity(payload)
    return True


def _validate_xlsx_archive(raw: bytes) -> frozenset[str]:
    try:
        archive = ZipFile(io.BytesIO(raw))
    except (BadZipFile, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    total_bytes = 0
    checked_xml_entries: set[str] = set()
    try:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ZIP_ENTRIES:
            raise AttachmentPreprocessingError("xlsx_archive_too_large")
        for entry in entries:
            normalized_name = entry.filename.replace("\\", "/").casefold()
            if entry.flag_bits & 0x1:
                raise AttachmentPreprocessingError("xlsx_encrypted_unsupported")
            if normalized_name.endswith("vbaproject.bin"):
                raise AttachmentPreprocessingError("xlsx_macros_unsupported")
            if entry.file_size < 0 or entry.file_size > MAX_XLSX_ZIP_ENTRY_BYTES:
                raise AttachmentPreprocessingError("xlsx_archive_too_large")
            total_bytes += entry.file_size
            if total_bytes > MAX_XLSX_ZIP_TOTAL_BYTES:
                raise AttachmentPreprocessingError("xlsx_archive_too_large")
            if entry.compress_size == 0:
                if entry.file_size > 0:
                    raise AttachmentPreprocessingError("xlsx_archive_too_large")
            elif entry.file_size / entry.compress_size > MAX_XLSX_ZIP_COMPRESSION_RATIO:
                raise AttachmentPreprocessingError("xlsx_archive_too_large")
            if _check_classified_xml_entry(archive, entry):
                checked_xml_entries.add(entry.filename)
    finally:
        archive.close()
    return frozenset(checked_xml_entries)


@dataclass(frozen=True)
class _WorksheetXmlPreflight:
    sheet_name: str
    archive_path: str
    stored_cells: int
    observed_max_row: int
    observed_max_column: int
    reported_max_row: int | None
    reported_max_column: int | None


def _resolved_package_target(target: object, *, source_part: str) -> str:
    if not isinstance(target, str) or not target or "\x00" in target or "\\" in target:
        raise AttachmentPreprocessingError("xlsx_parse_failed")
    normalized = target
    parts: list[str] = [] if normalized.startswith("/") else source_part.split("/")[:-1]
    for part in normalized.split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            if not parts:
                raise AttachmentPreprocessingError("xlsx_parse_failed")
            parts.pop()
            continue
        if ":" in part:
            raise AttachmentPreprocessingError("xlsx_parse_failed")
        parts.append(part)
    if not parts:
        raise AttachmentPreprocessingError("xlsx_parse_failed")
    return "/".join(parts)


def _relationships_part_for(source_part: str) -> str:
    parts = source_part.split("/")
    if len(parts) < 2:
        raise AttachmentPreprocessingError("xlsx_workbook_part_unsupported")
    return "/".join((*parts[:-1], "_rels", f"{parts[-1]}.rels"))


def _archive_part_identity(value: object) -> str:
    if not isinstance(value, str) or not value or "\x00" in value or "\\" in value:
        raise AttachmentPreprocessingError("xlsx_parse_failed")
    normalized = value
    if normalized.startswith("/"):
        normalized = normalized[1:]
    parts = normalized.split("/")
    if not parts or any(part in ("", ".", "..") or ":" in part for part in parts):
        raise AttachmentPreprocessingError("xlsx_parse_failed")
    return "/".join(parts)


def _xml_local_name(tag: object) -> str:
    return str(tag).rsplit("}", 1)[-1]


@dataclass(frozen=True)
class _ExactXmlCollectionItem:
    tag: str
    attributes: dict[str, str]


def _exact_collection_items(
    stream: Any,
    *,
    root_tag: str,
    collection_tag: str | None,
    child_tags: frozenset[str],
    error_code: str,
) -> list[_ExactXmlCollectionItem]:
    items: list[_ExactXmlCollectionItem] = []
    depth = 0
    root_seen = False
    root_closed = False
    collection_seen = False
    collection_depth: int | None = None
    active_child_depth: int | None = None
    collection_local_name = _xml_local_name(collection_tag) if collection_tag else None
    for event, element in ElementTree.iterparse(stream, events=("start", "end")):
        if event == "start":
            depth += 1
            if depth == 1:
                if root_seen or element.tag != root_tag:
                    raise AttachmentPreprocessingError(error_code)
                root_seen = True
                if collection_tag is None:
                    collection_seen = True
                    collection_depth = depth
            elif (
                collection_tag is not None
                and _xml_local_name(element.tag) == collection_local_name
            ):
                if element.tag != collection_tag or depth != 2 or collection_seen:
                    raise AttachmentPreprocessingError(error_code)
                collection_seen = True
                collection_depth = depth

            if collection_depth is not None and depth == collection_depth + 1:
                if element.tag not in child_tags or active_child_depth is not None:
                    raise AttachmentPreprocessingError(error_code)
                active_child_depth = depth
            elif active_child_depth is not None and depth > active_child_depth:
                raise AttachmentPreprocessingError(error_code)
        else:
            if active_child_depth is not None and depth == active_child_depth:
                if element.tag not in child_tags:
                    raise AttachmentPreprocessingError(error_code)
                items.append(
                    _ExactXmlCollectionItem(
                        tag=str(element.tag),
                        attributes=dict(element.attrib),
                    )
                )
                active_child_depth = None
            if collection_depth is not None and depth == collection_depth:
                if collection_tag is not None and element.tag != collection_tag:
                    raise AttachmentPreprocessingError(error_code)
                collection_depth = None
            if depth == 1:
                if element.tag != root_tag:
                    raise AttachmentPreprocessingError(error_code)
                root_closed = True
            element.clear()
            depth -= 1
    if (
        depth != 0
        or not root_seen
        or not root_closed
        or not collection_seen
        or collection_depth is not None
        or active_child_depth is not None
    ):
        raise AttachmentPreprocessingError(error_code)
    return items


def _canonical_workbook_part(archive: ZipFile) -> tuple[str, str]:
    try:
        with archive.open("[Content_Types].xml", "r") as stream:
            manifest_items = _exact_collection_items(
                stream,
                root_tag=_CONTENT_TYPES_ROOT_TAG,
                collection_tag=None,
                child_tags=frozenset({_CONTENT_TYPES_DEFAULT_TAG, _CONTENT_TYPES_OVERRIDE_TAG}),
                error_code="xlsx_content_types_structure_unsupported",
            )
        defaults: set[str] = set()
        override_parts: set[str] = set()
        workbook_defaults: list[tuple[str, str]] = []
        workbook_overrides: list[tuple[str, str]] = []
        for item in manifest_items:
            attributes = item.attributes
            if item.tag == _CONTENT_TYPES_DEFAULT_TAG:
                if set(attributes) != {"Extension", "ContentType"}:
                    raise AttachmentPreprocessingError("xlsx_content_types_structure_unsupported")
                extension = attributes.get("Extension")
                content_type = attributes.get("ContentType")
                if (
                    not isinstance(extension, str)
                    or not extension
                    or "/" in extension
                    or "\\" in extension
                    or extension.casefold() in defaults
                    or not isinstance(content_type, str)
                    or not content_type
                ):
                    raise AttachmentPreprocessingError("xlsx_content_types_structure_unsupported")
                defaults.add(extension.casefold())
                if content_type in _OPENPYXL_WORKBOOK_CONTENT_TYPES:
                    workbook_defaults.append((extension, content_type))
                continue

            if set(attributes) != {"PartName", "ContentType"}:
                raise AttachmentPreprocessingError("xlsx_content_types_structure_unsupported")
            raw_part_name = attributes.get("PartName")
            if not isinstance(raw_part_name, str) or not raw_part_name.startswith("/"):
                raise AttachmentPreprocessingError("xlsx_content_types_structure_unsupported")
            part_name = _archive_part_identity(raw_part_name)
            content_type = attributes.get("ContentType")
            if part_name in override_parts or not isinstance(content_type, str) or not content_type:
                raise AttachmentPreprocessingError("xlsx_content_types_structure_unsupported")
            override_parts.add(part_name)
            if content_type in _OPENPYXL_WORKBOOK_CONTENT_TYPES:
                workbook_overrides.append((raw_part_name, content_type))

        supported = (
            workbook_overrides
            == [(f"/{_CANONICAL_WORKBOOK_PART}", _XLSX_WORKBOOK_CONTENT_TYPE)]
            and not workbook_defaults
        ) or (
            not workbook_overrides
            and workbook_defaults == [("xml", _XLSX_WORKBOOK_CONTENT_TYPE)]
        )
        if not supported:
            raise AttachmentPreprocessingError("xlsx_workbook_part_unsupported")
    except AttachmentPreprocessingError:
        raise
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError, RuntimeError, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    return _CANONICAL_WORKBOOK_PART, _relationships_part_for(_CANONICAL_WORKBOOK_PART)


def _selected_worksheet_entries(
    archive: ZipFile,
    *,
    workbook_part: str,
    relationships_part: str,
) -> list[tuple[str, str]]:
    try:
        with archive.open(relationships_part, "r") as stream:
            relationship_items = _exact_collection_items(
                stream,
                root_tag=_PACKAGE_RELATIONSHIPS_ROOT_TAG,
                collection_tag=None,
                child_tags=frozenset({_PACKAGE_RELATIONSHIP_TAG}),
                error_code="xlsx_relationship_structure_unsupported",
            )

        worksheet_relationships: dict[str, str] = {}
        relationship_ids: set[str] = set()
        worksheet_targets: set[str] = set()
        relationship_type_prefix = f"{_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE}/"
        for item in relationship_items:
            attributes = item.attributes
            relationship_id = attributes.get("Id")
            relationship_type = attributes.get("Type")
            target_mode = str(attributes.get("TargetMode") or "")
            if (
                not isinstance(relationship_id, str)
                or not relationship_id
                or "\x00" in relationship_id
                or relationship_id in relationship_ids
                or not isinstance(relationship_type, str)
                or not relationship_type.startswith(relationship_type_prefix)
                or target_mode not in ("", "Internal")
            ):
                raise AttachmentPreprocessingError("xlsx_relationship_structure_unsupported")
            relationship_ids.add(relationship_id)
            target = _resolved_package_target(attributes.get("Target"), source_part=workbook_part)
            if relationship_type == _WORKSHEET_RELATIONSHIP_TYPE:
                if target in worksheet_targets:
                    raise AttachmentPreprocessingError("xlsx_relationship_structure_unsupported")
                worksheet_targets.add(target)
                worksheet_relationships[relationship_id] = target

        with archive.open(workbook_part, "r") as stream:
            sheet_items = _exact_collection_items(
                stream,
                root_tag=_WORKBOOK_ROOT_TAG,
                collection_tag=_WORKBOOK_SHEETS_TAG,
                child_tags=frozenset({_WORKBOOK_SHEET_TAG}),
                error_code="xlsx_workbook_structure_unsupported",
            )

        entries: list[tuple[str, str]] = []
        sheet_names: set[str] = set()
        sheet_ids: set[int] = set()
        worksheet_paths: set[str] = set()
        for item in sheet_items:
            attributes = item.attributes
            sheet_name = attributes.get("name")
            raw_sheet_id = attributes.get("sheetId")
            relationship_id = attributes.get(_WORKBOOK_SHEET_RELATIONSHIP_ID)
            archive_path = worksheet_relationships.get(str(relationship_id or ""))
            try:
                if not isinstance(raw_sheet_id, str):
                    raise ValueError
                sheet_id = int(raw_sheet_id)
            except (TypeError, ValueError) as exc:
                raise AttachmentPreprocessingError(
                    "xlsx_workbook_structure_unsupported"
                ) from exc
            if (
                not isinstance(sheet_name, str)
                or not sheet_name
                or sheet_name in sheet_names
                or sheet_id <= 0
                or sheet_id in sheet_ids
                or archive_path is None
                or archive_path in worksheet_paths
            ):
                raise AttachmentPreprocessingError("xlsx_workbook_structure_unsupported")
            sheet_names.add(sheet_name)
            sheet_ids.add(sheet_id)
            worksheet_paths.add(archive_path)
            entries.append((sheet_name, archive_path))
    except AttachmentPreprocessingError:
        raise
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError, RuntimeError, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    return entries


def _cell_reference_coordinates(value: object) -> tuple[int, int]:
    if not isinstance(value, str) or not value or len(value) > 10:
        raise ValueError("invalid cell reference")
    letter_end = 0
    while letter_end < len(value) and value[letter_end].isalpha():
        letter_end += 1
    letters = value[:letter_end].upper()
    digits = value[letter_end:]
    if not (1 <= len(letters) <= 3) or not digits or not digits.isascii() or not digits.isdigit():
        raise ValueError("invalid cell reference")
    column = 0
    for character in letters:
        if character < "A" or character > "Z":
            raise ValueError("invalid cell reference")
        column = column * 26 + ord(character) - ord("A") + 1
    row = int(digits)
    if not (1 <= row <= 1_048_576) or not (1 <= column <= 16_384):
        raise ValueError("invalid cell reference")
    return row, column


def _dimension_bounds(value: object) -> tuple[int | None, int | None]:
    if not isinstance(value, str):
        return None, None
    references = value.split(":")
    if len(references) not in (1, 2):
        return None, None
    try:
        coordinates = [_cell_reference_coordinates(reference) for reference in references]
    except ValueError:
        return None, None
    if len(coordinates) == 1:
        return coordinates[0]
    (start_row, start_column), (end_row, end_column) = coordinates
    if start_row > end_row or start_column > end_column:
        return None, None
    return end_row, end_column


def _worksheet_xml_preflight(
    archive: ZipFile,
    *,
    sheet_name: str,
    archive_path: str,
    cells_seen: int,
    content_security_checked: bool,
) -> tuple[_WorksheetXmlPreflight, int]:
    stored_cells = 0
    observed_max_row = 0
    observed_max_column = 0
    reported_max_row: int | None = None
    reported_max_column: int | None = None
    dimension_seen = False
    current_row: int | None = None
    current_column = 0
    last_row = 0
    worksheet_prefix = f"{{{_SPREADSHEET_XML_NAMESPACE}}}"
    dimension_tag = f"{worksheet_prefix}dimension"
    row_tag = f"{worksheet_prefix}row"
    cell_tag = f"{worksheet_prefix}c"
    depth = 0
    active_row_depth: int | None = None
    try:
        if content_security_checked:
            stream_context = archive.open(archive_path, "r")
        else:
            payload = archive.read(archive_path)
            _assert_xml_payload_has_no_dtd_or_entity(payload)
            stream_context = io.BytesIO(payload)
        with stream_context as stream:
            for event, element in ElementTree.iterparse(stream, events=("start", "end")):
                if event == "start":
                    depth += 1
                if (
                    event == "start"
                    and active_row_depth is not None
                    and depth == active_row_depth + 1
                ):
                    if element.tag != cell_tag:
                        raise AttachmentPreprocessingError("xlsx_worksheet_structure_unsupported")
                    stored_cells += 1
                    cells_seen += 1
                    if cells_seen > MAX_XLSX_CELLS:
                        raise AttachmentPreprocessingError("xlsx_cell_limit_exceeded")
                    raw_coordinate = element.attrib.get("r")
                    if raw_coordinate is None:
                        if current_row is None:
                            raise ValueError("cell outside row")
                        row, column = current_row, current_column + 1
                    else:
                        row, column = _cell_reference_coordinates(raw_coordinate)
                    if current_row is None or row != current_row or column <= current_column:
                        raise ValueError("invalid cell order")
                    current_column = column
                    observed_max_row = max(observed_max_row, row)
                    observed_max_column = max(observed_max_column, column)
                elif event == "start" and element.tag == dimension_tag:
                    if dimension_seen:
                        reported_max_row = None
                        reported_max_column = None
                    else:
                        reported_max_row, reported_max_column = _dimension_bounds(element.attrib.get("ref"))
                    dimension_seen = True
                elif event == "start" and element.tag == row_tag:
                    if active_row_depth is not None:
                        raise AttachmentPreprocessingError("xlsx_worksheet_structure_unsupported")
                    active_row_depth = depth
                    raw_row = element.attrib.get("r")
                    if raw_row is None:
                        current_row = last_row + 1
                    elif not raw_row.isascii() or not raw_row.isdigit():
                        raise ValueError("invalid row reference")
                    else:
                        current_row = int(raw_row)
                    if current_row < 1 or current_row > 1_048_576 or current_row <= last_row:
                        raise ValueError("invalid row reference")
                    last_row = current_row
                    current_column = 0
                elif event == "end" and element.tag == row_tag and depth == active_row_depth:
                    active_row_depth = None
                    current_row = None
                if event == "end":
                    element.clear()
                    depth -= 1
    except AttachmentPreprocessingError:
        raise
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError, RuntimeError, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    return (
        _WorksheetXmlPreflight(
            sheet_name=sheet_name,
            archive_path=archive_path,
            stored_cells=stored_cells,
            observed_max_row=observed_max_row,
            observed_max_column=observed_max_column,
            reported_max_row=reported_max_row,
            reported_max_column=reported_max_column,
        ),
        cells_seen,
    )


def _preflight_xlsx_worksheets(
    raw: bytes,
    *,
    content_security_checked_entries: frozenset[str],
) -> tuple[list[_WorksheetXmlPreflight], int, int]:
    try:
        archive = ZipFile(io.BytesIO(raw))
    except (BadZipFile, ValueError) as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    facts: list[_WorksheetXmlPreflight] = []
    cells_seen = 0
    try:
        workbook_part, relationships_part = _canonical_workbook_part(archive)
        worksheet_entries = _selected_worksheet_entries(
            archive,
            workbook_part=workbook_part,
            relationships_part=relationships_part,
        )
        if len({archive_path for _sheet_name, archive_path in worksheet_entries}) != len(worksheet_entries):
            raise AttachmentPreprocessingError("xlsx_parse_failed")
        for sheet_name, archive_path in worksheet_entries:
            sheet_facts, cells_seen = _worksheet_xml_preflight(
                archive,
                sheet_name=sheet_name,
                archive_path=archive_path,
                cells_seen=cells_seen,
                content_security_checked=archive_path in content_security_checked_entries,
            )
            facts.append(sheet_facts)
    finally:
        archive.close()
    return facts, cells_seen, len(worksheet_entries)


def _bounded_cell_payload(cell: Any) -> tuple[dict[str, Any] | None, bool]:
    value = getattr(cell, "value", None)
    if value is None:
        return None, False
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
        kind = "datetime"
    elif isinstance(value, bool):
        return {"column": int(cell.column), "kind": "boolean", "value": value}, False
    elif isinstance(value, (int, float)):
        return {"column": int(cell.column), "kind": "number", "value": value}, False
    else:
        text = str(value)
        kind = "formula" if getattr(cell, "data_type", "") == "f" or text.startswith("=") else "text"
    truncated = len(text) > MAX_XLSX_CELL_CHARS
    return {
        "column": int(cell.column),
        "kind": kind,
        "value": text[:MAX_XLSX_CELL_CHARS],
    }, truncated


def _prompt_content_within_caps(content: dict[str, Any]) -> bool:
    rendered = json.dumps(content, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return len(rendered) <= MAX_XLSX_PROMPT_CHARS and utf8_token_estimate(rendered) <= MAX_XLSX_PROMPT_TOKENS


def _reported_dimension_bound(value: object) -> int | None:
    if isinstance(value, bool) or not isinstance(value, int) or value < 1:
        return None
    return value


def parse_xlsx_attachment(
    *,
    path: Path,
    requirement: AttachmentParserRequirement,
) -> ParsedAttachmentContext:
    """Parse one broker-staged XLSX with deterministic, bounded read-only rules."""

    spec = parser_spec_for_attachment(
        file_name=requirement.file_name,
        content_type=requirement.content_type,
    )
    if not requirement.supported or spec is None:
        raise AttachmentPreprocessingError("attachment_parser_unsupported")
    if (
        requirement.parser_id != spec.parser_id
        or requirement.parser_version != spec.parser_version
        or requirement.max_bytes != spec.max_bytes
        or (
            requirement.extension not in spec.extensions
            and _normalized_content_type(requirement.content_type) not in spec.content_types
        )
    ):
        raise AttachmentPreprocessingError("attachment_preprocessing_contract_invalid")
    if path.name != requirement.file_name or path.suffix.casefold() != requirement.extension:
        raise AttachmentPreprocessingError("attachment_parser_staged_file_mismatch")
    try:
        raw = path.read_bytes()
    except OSError as exc:
        raise AttachmentPreprocessingError("attachment_parser_staged_file_invalid") from exc
    if len(raw) > requirement.max_bytes:
        raise AttachmentPreprocessingError("attachment_parser_file_too_large")
    actual_sha256 = hashlib.sha256(raw).hexdigest()
    if requirement.expected_byte_count is not None and len(raw) != requirement.expected_byte_count:
        raise AttachmentPreprocessingError("attachment_parser_staged_file_mismatch")
    if requirement.expected_sha256 is not None and actual_sha256 != requirement.expected_sha256:
        raise AttachmentPreprocessingError("attachment_parser_staged_file_mismatch")
    content_security_checked_entries = _validate_xlsx_archive(raw)
    worksheet_preflight, stored_cells, preflight_sheet_count = _preflight_xlsx_worksheets(
        raw,
        content_security_checked_entries=content_security_checked_entries,
    )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(raw),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc

    content: dict[str, Any] = {
        "schema_version": ATTACHMENT_CONTEXT_SCHEMA_VERSION,
        "file_id": requirement.file_id,
        "workbook": {"sheets": []},
    }
    cells_examined = stored_cells
    nonempty_cells = 0
    rows_emitted = 0
    sheets_processed = 0
    truncated = False
    stop_all = False
    try:
        sheet_names = list(workbook.sheetnames)
        if len(sheet_names) != preflight_sheet_count:
            raise AttachmentPreprocessingError("xlsx_parse_failed")
        if [str(sheet_name) for sheet_name in sheet_names] != [
            facts.sheet_name for facts in worksheet_preflight
        ]:
            raise AttachmentPreprocessingError("xlsx_parse_failed")
        for sheet_index, sheet_name in enumerate(sheet_names):
            worksheet = workbook[sheet_name]
            actual_archive_path = _archive_part_identity(
                getattr(worksheet, "_worksheet_path", None)
            )
            if actual_archive_path != worksheet_preflight[sheet_index].archive_path:
                raise AttachmentPreprocessingError("xlsx_parse_failed")
        selected_sheet_names = sheet_names[:MAX_XLSX_SHEETS]
        if preflight_sheet_count > MAX_XLSX_SHEETS:
            truncated = True
        for sheet_index, sheet_name in enumerate(selected_sheet_names):
            worksheet = workbook[sheet_name]
            xml_facts = worksheet_preflight[sheet_index]
            max_row = xml_facts.reported_max_row
            max_column = xml_facts.reported_max_column
            if len(str(sheet_name)) > MAX_XLSX_CELL_CHARS:
                truncated = True
            if (
                max_row is None
                or max_column is None
                or max_row < xml_facts.observed_max_row
                or max_column < xml_facts.observed_max_column
            ):
                max_row = None
                max_column = None
                truncated = True
            elif max_row > MAX_XLSX_ROWS_PER_SHEET or max_column > MAX_XLSX_COLUMNS_PER_SHEET:
                truncated = True
            if (
                xml_facts.observed_max_row > MAX_XLSX_ROWS_PER_SHEET
                or xml_facts.observed_max_column > MAX_XLSX_COLUMNS_PER_SHEET
            ):
                truncated = True
            sheet_payload: dict[str, Any] = {
                "name": str(sheet_name)[:MAX_XLSX_CELL_CHARS],
                "max_row": max_row,
                "max_column": max_column,
                "rows": [],
            }
            content["workbook"]["sheets"].append(sheet_payload)
            sheets_processed += 1
            if not _prompt_content_within_caps(content):
                content["workbook"]["sheets"].pop()
                sheets_processed -= 1
                truncated = True
                break
            for row_index, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    min_col=1,
                    max_row=MAX_XLSX_ROWS_PER_SHEET,
                    max_col=MAX_XLSX_COLUMNS_PER_SHEET,
                ),
                start=1,
            ):
                row_cells: list[dict[str, Any]] = []
                for cell in row:
                    if _reported_dimension_bound(getattr(cell, "column", None)) is None:
                        continue
                    cell_payload, cell_truncated = _bounded_cell_payload(cell)
                    truncated = truncated or cell_truncated
                    if cell_payload is not None:
                        nonempty_cells += 1
                        row_cells.append(cell_payload)
                if row_cells:
                    row_payload = {"row": row_index, "cells": row_cells}
                    sheet_payload["rows"].append(row_payload)
                    if not _prompt_content_within_caps(content):
                        sheet_payload["rows"].pop()
                        truncated = True
                        stop_all = True
                        break
                    rows_emitted += 1
                if stop_all:
                    break
            if stop_all:
                break
    except AttachmentPreprocessingError:
        raise
    except Exception as exc:
        raise AttachmentPreprocessingError("xlsx_parse_failed") from exc
    finally:
        workbook.close()

    content["workbook"]["sheet_count"] = len(sheet_names)
    content["workbook"]["truncated"] = truncated
    if not _prompt_content_within_caps(content):
        raise AttachmentPreprocessingError("attachment_parser_prompt_too_large")
    evidence = AttachmentParserEvidence(
        file_id=requirement.file_id,
        parser_id=spec.parser_id,
        parser_version=spec.parser_version,
        content_type=requirement.content_type,
        extension=requirement.extension,
        byte_count=len(raw),
        sha256=actual_sha256,
        sheet_count=len(sheet_names),
        sheets_processed=sheets_processed,
        cells_examined=cells_examined,
        nonempty_cells=nonempty_cells,
        rows_emitted=rows_emitted,
        truncated=truncated,
        status="parsed",
    )
    return ParsedAttachmentContext(evidence=evidence, content=content)


def validate_required_parser_evidence(
    *,
    requirements: list[AttachmentParserRequirement],
    evidence: object,
) -> tuple[bool, str]:
    """Require one exact positive evidence record for every supported workbook."""

    if any(not requirement.supported for requirement in requirements):
        return False, "attachment_parser_unsupported"
    if not requirements:
        return True, ""
    if not isinstance(evidence, list):
        return False, "attachment_parser_evidence_missing"
    parsed_by_file: dict[str, AttachmentParserEvidence] = {}
    for raw in evidence:
        try:
            item = AttachmentParserEvidence.model_validate(raw)
        except Exception:
            return False, "attachment_parser_evidence_invalid"
        if item.file_id in parsed_by_file:
            return False, "attachment_parser_evidence_invalid"
        parsed_by_file[item.file_id] = item
    for requirement in requirements:
        item = parsed_by_file.get(requirement.file_id)
        if item is None:
            return False, "attachment_parser_evidence_missing"
        if (
            item.parser_id != requirement.parser_id
            or item.parser_version != requirement.parser_version
            or item.content_type != requirement.content_type
            or item.extension != requirement.extension
            or item.byte_count > requirement.max_bytes
            or (
                requirement.expected_byte_count is not None
                and item.byte_count != requirement.expected_byte_count
            )
            or (
                requirement.expected_sha256 is not None
                and item.sha256 != requirement.expected_sha256
            )
            or item.sheets_processed > item.sheet_count
            or item.nonempty_cells > item.cells_examined
        ):
            return False, "attachment_parser_evidence_mismatch"
    return True, ""
