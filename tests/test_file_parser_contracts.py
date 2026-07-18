import hashlib
import io
import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import Workbook, load_workbook

from app.context_manifest import utf8_token_estimate
from app.executors.claude_agent_sdk_runner import _attachment_context_data_message
from app.file_parser_contracts import (
    MAX_XLSX_CELL_CHARS,
    MAX_XLSX_CELLS,
    MAX_XLSX_COLUMNS_PER_SHEET,
    MAX_XLSX_FILE_BYTES,
    MAX_XLSX_PROMPT_CHARS,
    MAX_XLSX_PROMPT_TOKENS,
    MAX_XLSX_ROWS_PER_SHEET,
    MAX_XLSX_SHEETS,
    AttachmentPreprocessingError,
    MaterializedAttachmentFact,
    attachment_requirements_from_contract,
    build_attachment_preprocessing_contract,
    is_known_binary_workbook,
    parse_xlsx_attachment,
    validate_required_parser_evidence,
)


def _requirement(file_name: str = "book.xlsx"):
    contract = build_attachment_preprocessing_contract(
        file_ids=["file-a"],
        file_names=[file_name],
    )
    return attachment_requirements_from_contract(contract)[0]


def _write_workbook(path: Path, *, long: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "name"
    sheet["B1"] = "value"
    sheet["A2"] = "alpha"
    sheet["B2"] = "=1+2"
    if long:
        sheet["C2"] = "界" * (MAX_XLSX_CELL_CHARS + 10)
        for row in range(3, MAX_XLSX_ROWS_PER_SHEET + 3):
            sheet.cell(row=row, column=1, value=f"row-{row}")
    workbook.save(path)
    workbook.close()


def _remove_worksheet_dimension(path: Path) -> None:
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    worksheet_path = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == worksheet_path:
                root = ElementTree.fromstring(payload)
                dimension = root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension")
                assert dimension is not None
                root.remove(dimension)
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())
    with zipfile.ZipFile(path, "r") as archive:
        root = ElementTree.fromstring(archive.read(worksheet_path))
        assert root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension") is None


def _set_worksheet_dimension(path: Path, reference: str) -> None:
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    worksheet_path = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == worksheet_path:
                root = ElementTree.fromstring(payload)
                dimension = root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension")
                assert dimension is not None
                dimension.set("ref", reference)
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def _inject_worksheet_entity_declaration(path: Path, *, utf16: bool = False) -> None:
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    worksheet_path = "xl/worksheets/sheet1.xml"
    declaration = b'<!DOCTYPE worksheet [<!ENTITY unsafe "blocked">]>'
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == worksheet_path:
                insertion = payload.find(b"<worksheet")
                assert insertion >= 0
                payload = payload[:insertion] + declaration + payload[insertion:]
                if utf16:
                    payload = payload.decode("utf-8").encode("utf-16")
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def _write_dimensionless_validation_workbook(path: Path, *, overflow: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validation"
    rows = [
        ["Requirement", "Control", "Status", "Evidence"],
        ["GMP-VAL-002 Requirement", "Validate import", "Pass", "ACCEPT-XLSX-9472"],
        ["Owner", "Quality", "Reviewed", "Yes"],
        ["System", "AI Platform", "Mode", "Read only"],
        ["Parser", "openpyxl", "Formula", "Not executed"],
        ["Boundary", "Bounded", "External links", "Disabled"],
        ["Decision", "Accepted"],
    ]
    for row in rows:
        sheet.append(row)
    if overflow:
        sheet.cell(row=MAX_XLSX_ROWS_PER_SHEET + 1, column=1, value="ROW-101-MUST-BE-EXCLUDED")
        sheet.cell(row=1, column=MAX_XLSX_COLUMNS_PER_SHEET + 1, value="COL-33-MUST-BE-EXCLUDED")
    workbook.save(path)
    workbook.close()
    _remove_worksheet_dimension(path)


def _write_stored_cell_overflow_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "inside-a"
    sheet["B1"] = "inside-b"
    for column in range(1, MAX_XLSX_CELLS + 2):
        sheet.cell(row=MAX_XLSX_ROWS_PER_SHEET + 1, column=column, value=column)
    workbook.save(path)
    workbook.close()


def _write_exact_cell_limit_overflow_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "inside"
    for column in range(1, MAX_XLSX_CELLS + 1):
        sheet.cell(row=MAX_XLSX_ROWS_PER_SHEET + 1, column=column, value=column)
    workbook.save(path)
    workbook.close()


def _inject_foreign_relationship_id_decoy(path: Path) -> None:
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_relationships_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    content_types_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    markup_compatibility_namespace = (
        "http://schemas.openxmlformats.org/markup-compatibility/2006"
    )
    foreign_namespace = "urn:ai-platform:test:foreign"
    ElementTree.register_namespace("r", office_relationships_namespace)
    ElementTree.register_namespace("mc", markup_compatibility_namespace)
    ElementTree.register_namespace("foo", foreign_namespace)
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    decoy_path = "xl/worksheets/decoy.xml"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "xl/workbook.xml":
                root = ElementTree.fromstring(payload)
                root.set(f"{{{markup_compatibility_namespace}}}Ignorable", "foo")
                sheet = root.find(f".//{{{spreadsheet_namespace}}}sheet")
                assert sheet is not None
                real_relationship_id = sheet.attrib.pop(
                    f"{{{office_relationships_namespace}}}id"
                )
                sheet.set(f"{{{foreign_namespace}}}id", "rFake")
                sheet.set(
                    f"{{{office_relationships_namespace}}}id",
                    real_relationship_id,
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                root.insert(
                    0,
                    ElementTree.Element(
                        f"{{{package_relationships_namespace}}}Relationship",
                        {
                            "Id": "rFake",
                            "Type": f"{office_relationships_namespace}/worksheet",
                            "Target": f"/{decoy_path}",
                        },
                    ),
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                root.append(
                    ElementTree.Element(
                        f"{{{content_types_namespace}}}Override",
                        {
                            "PartName": f"/{decoy_path}",
                            "ContentType": (
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.worksheet+xml"
                            ),
                        },
                    )
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
        rewritten.writestr(
            decoy_path,
            (
                f'<worksheet xmlns="{spreadsheet_namespace}">'
                '<dimension ref="A1"/><sheetData><row r="1">'
                '<c r="A1" t="inlineStr"><is><t>decoy</t></is></c>'
                "</row></sheetData></worksheet>"
            ).encode("utf-8"),
        )
    path.write_bytes(output.getvalue())


def _inject_foreign_row_children(path: Path, *, count: int) -> None:
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    markup_compatibility_namespace = (
        "http://schemas.openxmlformats.org/markup-compatibility/2006"
    )
    foreign_namespace = "urn:ai-platform:test:foreign-row"
    ElementTree.register_namespace("mc", markup_compatibility_namespace)
    ElementTree.register_namespace("foo", foreign_namespace)
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    worksheet_path = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == worksheet_path:
                root = ElementTree.fromstring(payload)
                root.set(f"{{{markup_compatibility_namespace}}}Ignorable", "foo")
                row = root.find(f".//{{{spreadsheet_namespace}}}row")
                assert row is not None and len(row) == 1
                for index in range(count):
                    foreign_cell = ElementTree.SubElement(
                        row,
                        f"{{{foreign_namespace}}}cell",
                        {"sequence": str(index)},
                    )
                    value = ElementTree.SubElement(
                        foreign_cell,
                        f"{{{spreadsheet_namespace}}}v",
                    )
                    value.text = str(index)
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def _overflow_worksheet_xml() -> bytes:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "inside"
    for column in range(1, MAX_XLSX_CELLS + 1):
        sheet.cell(row=MAX_XLSX_ROWS_PER_SHEET + 1, column=column, value=column)
    workbook.save(buffer)
    workbook.close()
    with zipfile.ZipFile(io.BytesIO(buffer.getvalue()), "r") as archive:
        return archive.read("xl/worksheets/sheet1.xml")


def _inject_foreign_relationship_lookalike(
    path: Path,
    *,
    local_name: str = "Relationship",
) -> None:
    office_relationships_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    content_types_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    markup_compatibility_namespace = (
        "http://schemas.openxmlformats.org/markup-compatibility/2006"
    )
    foreign_namespace = "urn:ai-platform:test:foreign-relationship"
    ElementTree.register_namespace("mc", markup_compatibility_namespace)
    ElementTree.register_namespace("foo", foreign_namespace)
    real_path = "xl/worksheets/real.xml"
    real_xml = _overflow_worksheet_xml()
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                root.set(f"{{{markup_compatibility_namespace}}}Ignorable", "foo")
                exact_relationship = root.find(
                    f"{{{package_relationships_namespace}}}Relationship"
                )
                assert exact_relationship is not None
                root.append(
                    ElementTree.Element(
                        f"{{{foreign_namespace}}}{local_name}",
                        {
                            "Id": exact_relationship.attrib["Id"],
                            "Type": f"{office_relationships_namespace}/worksheet",
                            "Target": f"/{real_path}",
                        },
                    )
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                root.append(
                    ElementTree.Element(
                        f"{{{content_types_namespace}}}Override",
                        {
                            "PartName": f"/{real_path}",
                            "ContentType": (
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.worksheet+xml"
                            ),
                        },
                    )
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
        rewritten.writestr(real_path, real_xml)
    path.write_bytes(output.getvalue())


def _inject_foreign_sheet_lookalike(path: Path, *, local_name: str = "sheet") -> None:
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_relationships_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    markup_compatibility_namespace = (
        "http://schemas.openxmlformats.org/markup-compatibility/2006"
    )
    foreign_namespace = "urn:ai-platform:test:foreign-sheet"
    ElementTree.register_namespace("r", office_relationships_namespace)
    ElementTree.register_namespace("mc", markup_compatibility_namespace)
    ElementTree.register_namespace("foo", foreign_namespace)
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "xl/workbook.xml":
                root = ElementTree.fromstring(payload)
                root.set(f"{{{markup_compatibility_namespace}}}Ignorable", "foo")
                sheets = root.find(f"{{{spreadsheet_namespace}}}sheets")
                exact_sheet = root.find(f".//{{{spreadsheet_namespace}}}sheet")
                assert sheets is not None and exact_sheet is not None
                sheets.append(
                    ElementTree.Element(
                        f"{{{foreign_namespace}}}{local_name}",
                        {
                            "name": "Foreign",
                            "sheetId": "2",
                            f"{{{office_relationships_namespace}}}id": exact_sheet.attrib[
                                f"{{{office_relationships_namespace}}}id"
                            ],
                        },
                    )
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def _retarget_worksheet_to_entity_payload(path: Path) -> None:
    office_relationships_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    content_types_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    source_path = "xl/worksheets/sheet1.xml"
    target_path = "xl/worksheets/sheet1.payload"
    declaration = b'<!DOCTYPE worksheet [<!ENTITY injected "EXPANDED">]>'
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            output_name = entry.filename
            if entry.filename == source_path:
                insertion = payload.find(b"<worksheet")
                assert insertion >= 0 and b"<t>exact</t>" in payload
                payload = payload[:insertion] + declaration + payload[insertion:]
                payload = payload.replace(b"<t>exact</t>", b"<t>&injected;</t>", 1)
                output_name = target_path
            elif entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                worksheet_relationship = next(
                    child
                    for child in root
                    if child.tag
                    == f"{{{package_relationships_namespace}}}Relationship"
                    and child.attrib.get("Type")
                    == f"{office_relationships_namespace}/worksheet"
                )
                worksheet_relationship.set("Target", f"/{target_path}")
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                worksheet_override = next(
                    child
                    for child in root
                    if child.tag == f"{{{content_types_namespace}}}Override"
                    and child.attrib.get("PartName") == f"/{source_path}"
                )
                worksheet_override.set("PartName", f"/{target_path}")
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            if output_name == entry.filename:
                rewritten.writestr(entry, payload)
            else:
                rewritten.writestr(output_name, payload)
    path.write_bytes(output.getvalue())


def _corrupt_exact_collection(path: Path, case: str) -> None:
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    foreign_namespace = "urn:ai-platform:test:collection"
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if case == "wrong-relationships-root" and entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                root.tag = f"{{{foreign_namespace}}}Relationships"
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif case == "nested-relationship" and entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                relationship = root.find(
                    f"{{{package_relationships_namespace}}}Relationship"
                )
                assert relationship is not None
                relationship.append(ElementTree.Element(f"{{{foreign_namespace}}}nested"))
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif case == "duplicate-sheets" and entry.filename == "xl/workbook.xml":
                root = ElementTree.fromstring(payload)
                root.append(ElementTree.Element(f"{{{spreadsheet_namespace}}}sheets"))
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def _select_alternate_workbook_with_duplicate_relationships(path: Path) -> None:
    content_types_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    alternate_workbook_path = "alt/workbook.xml"
    alternate_relationships_path = "alt/_rels/workbook.xml.rels"
    decoy_path = "xl/worksheets/alternate-decoy.xml"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        workbook_payload = archive.read("xl/workbook.xml")
        relationships_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        worksheet_relationship = next(
            child
            for child in relationships_root
            if child.attrib.get("Type", "").endswith("/worksheet")
        )
        duplicate = ElementTree.Element(
            f"{{{package_relationships_namespace}}}Relationship",
            dict(worksheet_relationship.attrib),
        )
        duplicate.set("Target", f"/{decoy_path}")
        relationships_root.append(duplicate)
        alternate_relationships_payload = ElementTree.tostring(
            relationships_root,
            encoding="utf-8",
            xml_declaration=True,
        )
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                workbook_override = next(
                    child
                    for child in root
                    if child.tag == f"{{{content_types_namespace}}}Override"
                    and child.attrib.get("PartName") == "/xl/workbook.xml"
                )
                workbook_override.set("PartName", f"/{alternate_workbook_path}")
                root.append(
                    ElementTree.Element(
                        f"{{{content_types_namespace}}}Override",
                        {
                            "PartName": f"/{decoy_path}",
                            "ContentType": (
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.worksheet+xml"
                            ),
                        },
                    )
                )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
        rewritten.writestr(alternate_workbook_path, workbook_payload)
        rewritten.writestr(alternate_relationships_path, alternate_relationships_payload)
        rewritten.writestr(decoy_path, archive.read("xl/worksheets/sheet1.xml"))
    path.write_bytes(output.getvalue())


def _retarget_later_worksheet_to_entity_payload(path: Path, *, sheet_index: int) -> None:
    content_types_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    source_path = f"xl/worksheets/sheet{sheet_index}.xml"
    target_path = f"xl/worksheets/sheet{sheet_index}.payload"
    declaration = b'<!DOCTYPE worksheet [<!ENTITY later "LATER-EXPANDED">]>'
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            output_name = entry.filename
            if entry.filename == source_path:
                insertion = payload.find(b"<worksheet")
                expected = f"<t>sheet-{sheet_index}</t>".encode("utf-8")
                assert insertion >= 0 and expected in payload
                payload = payload[:insertion] + declaration + payload[insertion:]
                payload = payload.replace(expected, b"<t>&later;</t>", 1)
                output_name = target_path
            elif entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                relationship = next(
                    child
                    for child in root
                    if child.attrib.get("Target", "").endswith(f"/{source_path}")
                )
                relationship.set("Target", f"/{target_path}")
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            elif entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                override = next(
                    child
                    for child in root
                    if child.tag == f"{{{content_types_namespace}}}Override"
                    and child.attrib.get("PartName") == f"/{source_path}"
                )
                override.set("PartName", f"/{target_path}")
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            if output_name == entry.filename:
                rewritten.writestr(entry, payload)
            else:
                rewritten.writestr(output_name, payload)
    path.write_bytes(output.getvalue())


def _inject_unreferenced_duplicate_worksheet_target(path: Path) -> None:
    package_relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(payload)
                worksheet_relationship = next(
                    child
                    for child in root
                    if child.attrib.get("Type", "").endswith("/worksheet")
                )
                duplicate = ElementTree.Element(
                    f"{{{package_relationships_namespace}}}Relationship",
                    dict(worksheet_relationship.attrib),
                )
                duplicate.set("Id", "rUnreferenced")
                root.append(duplicate)
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())


def test_xlsx_parser_emits_bounded_typed_content_and_positive_evidence(tmp_path):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    evidence = parsed.evidence
    assert evidence.status == "parsed"
    assert evidence.file_id == "file-a"
    assert evidence.parser_id == "ai-platform.xlsx.openpyxl"
    assert evidence.byte_count == path.stat().st_size
    assert evidence.sha256 == hashlib.sha256(path.read_bytes()).hexdigest()
    assert evidence.sheet_count == 1
    assert evidence.cells_examined >= 4
    assert evidence.truncated is False
    formula = parsed.content["workbook"]["sheets"][0]["rows"][1]["cells"][1]
    assert formula == {"column": 2, "kind": "formula", "value": "=1+2"}

    data_message = _attachment_context_data_message([parsed])
    assert '"message_kind":"platform_typed_attachment_data"' in data_message
    assert '"kind":"formula"' in data_message


def test_xlsx_parser_reports_deterministic_truncation(tmp_path):
    path = tmp_path / "book.xlsx"
    _write_workbook(path, long=True)

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    assert parsed.evidence.truncated is True
    assert parsed.content["workbook"]["truncated"] is True
    first_data_row = parsed.content["workbook"]["sheets"][0]["rows"][1]
    assert len(first_data_row["cells"][2]["value"]) == MAX_XLSX_CELL_CHARS


def test_xlsx_parser_reads_dimensionless_workbook_with_positive_evidence(tmp_path):
    path = tmp_path / "book.xlsx"
    _write_dimensionless_validation_workbook(path)

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    sheet = parsed.content["workbook"]["sheets"][0]
    rendered = json.dumps(parsed.content, ensure_ascii=False, sort_keys=True)
    assert sheet["name"] == "Validation"
    assert sheet["max_row"] is None
    assert sheet["max_column"] is None
    assert "GMP-VAL-002 Requirement" in rendered
    assert "ACCEPT-XLSX-9472" in rendered
    assert parsed.evidence.status == "parsed"
    assert parsed.evidence.nonempty_cells >= 26
    assert parsed.evidence.rows_emitted == 7
    assert parsed.evidence.truncated is True


def test_xlsx_parser_bounds_dimensionless_row_column_cell_and_prompt_content(tmp_path):
    path = tmp_path / "book.xlsx"
    _write_dimensionless_validation_workbook(path, overflow=True)

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    rendered = json.dumps(parsed.content, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    assert "ROW-101-MUST-BE-EXCLUDED" not in rendered
    assert "COL-33-MUST-BE-EXCLUDED" not in rendered
    assert parsed.evidence.cells_examined <= MAX_XLSX_CELLS
    assert len(rendered) <= MAX_XLSX_PROMPT_CHARS
    assert utf8_token_estimate(rendered) <= MAX_XLSX_PROMPT_TOKENS
    assert parsed.evidence.truncated is True
    assert parsed.content["workbook"]["truncated"] is True


def test_xlsx_parser_rejects_stored_cell_overflow_before_openpyxl_load(tmp_path, monkeypatch):
    path = tmp_path / "book.xlsx"
    _write_stored_cell_overflow_workbook(path)
    assert path.stat().st_size <= MAX_XLSX_FILE_BYTES
    with zipfile.ZipFile(path, "r") as archive:
        assert archive.read("xl/worksheets/sheet1.xml").count(b"<c ") == MAX_XLSX_CELLS + 3

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("openpyxl must not load content after the XML cell limit is exceeded")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_cell_limit_exceeded"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_marks_forged_low_dimension_unreliable(tmp_path):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "first"
    sheet["B2"] = "second"
    workbook.save(path)
    workbook.close()
    _set_worksheet_dimension(path, "A1")

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    sheet_content = parsed.content["workbook"]["sheets"][0]
    rendered = json.dumps(sheet_content, ensure_ascii=False, sort_keys=True)
    assert "first" in rendered
    assert "second" in rendered
    assert sheet_content["max_row"] is None
    assert sheet_content["max_column"] is None
    assert parsed.evidence.cells_examined == 2
    assert parsed.evidence.truncated is True


@pytest.mark.parametrize(
    ("reference", "expected_row", "expected_column", "expected_truncated"),
    [
        ("B2:A1", None, None, True),
        ("A2:B1", None, None, True),
        ("B1:A2", None, None, True),
        ("A1:B2", 2, 2, False),
    ],
    ids=["both-reversed", "row-reversed", "column-reversed", "valid"],
)
def test_xlsx_parser_validates_dimension_range_direction(
    tmp_path,
    reference,
    expected_row,
    expected_column,
    expected_truncated,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "stored"
    workbook.save(path)
    workbook.close()
    _set_worksheet_dimension(path, reference)

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    sheet_content = parsed.content["workbook"]["sheets"][0]
    assert sheet_content["max_row"] == expected_row
    assert sheet_content["max_column"] == expected_column
    assert parsed.evidence.truncated is expected_truncated


def test_xlsx_parser_ignores_foreign_id_and_rejects_real_sheet_overflow_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    _write_exact_cell_limit_overflow_workbook(path)
    _inject_foreign_relationship_id_decoy(path)
    assert path.stat().st_size <= MAX_XLSX_FILE_BYTES
    with zipfile.ZipFile(path, "r") as archive:
        assert archive.read("xl/worksheets/sheet1.xml").count(b"<c ") == MAX_XLSX_CELLS + 1
        assert archive.read("xl/worksheets/decoy.xml").count(b"<c ") == 1
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        workbook_sheet = workbook_root.find(
            ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
        )
        assert workbook_sheet is not None
        attribute_names = list(workbook_sheet.attrib)
        foreign_id = "{urn:ai-platform:test:foreign}id"
        real_id = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        assert attribute_names.index(foreign_id) < attribute_names.index(real_id)
    workbook_probe = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    assert workbook_probe.sheetnames == ["Sheet"]
    assert workbook_probe["Sheet"]._worksheet_path == "xl/worksheets/sheet1.xml"
    workbook_probe.close()

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("real worksheet overflow must fail before openpyxl loads the decoy workbook")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_cell_limit_exceeded"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_openpyxl_worksheet_part_mismatch(tmp_path, monkeypatch):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)

    class FakeWorksheet:
        _worksheet_path = "xl/worksheets/decoy.xml"

    class FakeWorkbook:
        sheetnames = ["Data"]
        closed = False

        def __getitem__(self, _sheet_name):
            return FakeWorksheet()

        def close(self):
            self.closed = True

    fake_workbook = FakeWorkbook()
    monkeypatch.setattr("openpyxl.load_workbook", lambda *_args, **_kwargs: fake_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_parse_failed"):
        parse_xlsx_attachment(path=path, requirement=_requirement())
    assert fake_workbook.closed is True


def test_xlsx_parser_rejects_foreign_row_children_before_openpyxl_load(tmp_path, monkeypatch):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "exact"
    workbook.save(path)
    workbook.close()
    _inject_foreign_row_children(path, count=MAX_XLSX_CELLS + 1)
    with zipfile.ZipFile(path, "r") as archive:
        root = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        row = root.find(
            ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"
        )
        assert row is not None
        assert len(row) == MAX_XLSX_CELLS + 2
        foreign_children = [
            child
            for child in row
            if child.tag == "{urn:ai-platform:test:foreign-row}cell"
        ]
        assert len(foreign_children) == MAX_XLSX_CELLS + 1
        assert foreign_children[0].find(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v"
        ).text == "0"

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("foreign row children must fail before openpyxl parses them")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_worksheet_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_foreign_relationship_lookalike_before_openpyxl_load(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "decoy"
    workbook.save(path)
    workbook.close()
    _inject_foreign_relationship_lookalike(path)
    assert path.stat().st_size <= MAX_XLSX_FILE_BYTES
    with zipfile.ZipFile(path, "r") as archive:
        assert archive.read("xl/worksheets/real.xml").count(b"<c ") == MAX_XLSX_CELLS + 1
        relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        relationship_type = (
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
        )
        exact_relationship = next(
            child
            for child in relationships
            if child.tag
            == "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
            and child.attrib.get("Type") == relationship_type
        )
        foreign_relationship = next(
            child
            for child in relationships
            if child.tag == "{urn:ai-platform:test:foreign-relationship}Relationship"
        )
        assert foreign_relationship.attrib["Id"] == exact_relationship.attrib["Id"]
        assert foreign_relationship.attrib["Target"] != exact_relationship.attrib["Target"]
        assert list(relationships).index(foreign_relationship) > list(relationships).index(
            exact_relationship
        )

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("foreign Relationship must fail before openpyxl redirects the part")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_relationship_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_foreign_sheet_lookalike_before_openpyxl_load(tmp_path, monkeypatch):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "exact"
    workbook.save(path)
    workbook.close()
    _inject_foreign_sheet_lookalike(path)
    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook_root.find(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets"
        )
        assert sheets is not None
        exact_sheet = next(
            child
            for child in sheets
            if child.tag
            == "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
        )
        foreign_sheet = next(
            child
            for child in sheets
            if child.tag == "{urn:ai-platform:test:foreign-sheet}sheet"
        )
        assert list(sheets).index(foreign_sheet) > list(sheets).index(exact_sheet)

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("foreign sheet must fail before openpyxl deserializes it")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_workbook_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_checks_selected_arbitrary_extension_payload_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "exact"
    workbook.save(path)
    workbook.close()
    _retarget_worksheet_to_entity_payload(path)
    with zipfile.ZipFile(path, "r") as archive:
        assert "xl/worksheets/sheet1.xml" not in archive.namelist()
        payload = archive.read("xl/worksheets/sheet1.payload")
        assert b"<!DOCTYPE" in payload
        assert b"&injected;" in payload

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("arbitrary-extension worksheet entities must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_xml_entities_unsupported"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_arbitrary_relationship_child_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "decoy"
    workbook.save(path)
    workbook.close()
    _inject_foreign_relationship_lookalike(path, local_name="redirect")
    with zipfile.ZipFile(path, "r") as archive:
        relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        assert any(
            child.tag == "{urn:ai-platform:test:foreign-relationship}redirect"
            for child in relationships
        )
        assert archive.read("xl/worksheets/real.xml").count(b"<c ") == MAX_XLSX_CELLS + 1

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("arbitrary Relationship child must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_relationship_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_arbitrary_seventeenth_sheet_child_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "sheet-1"
    for index in range(2, MAX_XLSX_SHEETS + 1):
        workbook.create_sheet(title=f"Sheet{index}")["A1"] = f"sheet-{index}"
    workbook.save(path)
    workbook.close()
    _inject_foreign_sheet_lookalike(path, local_name="item")
    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook_root.find(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets"
        )
        assert sheets is not None and len(sheets) == MAX_XLSX_SHEETS + 1
        assert sheets[-1].tag == "{urn:ai-platform:test:foreign-sheet}item"

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("arbitrary seventeenth sheet child must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_workbook_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


@pytest.mark.parametrize(
    ("case", "expected_code"),
    [
        ("wrong-relationships-root", "xlsx_relationship_structure_unsupported"),
        ("nested-relationship", "xlsx_relationship_structure_unsupported"),
        ("duplicate-sheets", "xlsx_workbook_structure_unsupported"),
    ],
)
def test_xlsx_parser_rejects_malformed_exact_collections_before_openpyxl(
    tmp_path,
    monkeypatch,
    case,
    expected_code,
):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)
    _corrupt_exact_collection(path, case)

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("malformed exact collection must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match=expected_code):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_preserves_total_sheet_count_beyond_selected_limit(tmp_path):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "sheet-1"
    for index in range(2, MAX_XLSX_SHEETS + 2):
        workbook.create_sheet(title=f"Sheet{index}")["A1"] = f"sheet-{index}"
    workbook.save(path)
    workbook.close()

    parsed = parse_xlsx_attachment(path=path, requirement=_requirement())

    assert parsed.evidence.sheet_count == MAX_XLSX_SHEETS + 1
    assert parsed.evidence.sheets_processed == MAX_XLSX_SHEETS
    assert parsed.evidence.truncated is True
    assert len(parsed.content["workbook"]["sheets"]) == MAX_XLSX_SHEETS


def test_xlsx_parser_rejects_alternate_manifest_workbook_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "canonical-decoy"
    workbook.save(path)
    workbook.close()
    _select_alternate_workbook_with_duplicate_relationships(path)
    with zipfile.ZipFile(path, "r") as archive:
        content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
        workbook_override = next(
            child
            for child in content_types
            if child.attrib.get("ContentType")
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        )
        assert workbook_override.attrib["PartName"] == "/alt/workbook.xml"
        alternate_relationships = ElementTree.fromstring(
            archive.read("alt/_rels/workbook.xml.rels")
        )
        relationship_ids = [child.attrib["Id"] for child in alternate_relationships]
        assert len(relationship_ids) != len(set(relationship_ids))

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("alternate manifest workbook must fail before openpyxl selection")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_workbook_part_unsupported"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_arbitrary_content_types_child_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)
    source = io.BytesIO(path.read_bytes())
    output = io.BytesIO()
    foreign_namespace = "urn:ai-platform:test:content-types"
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(output, "w") as rewritten:
        for entry in archive.infolist():
            payload = archive.read(entry.filename)
            if entry.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(payload)
                root.append(ElementTree.Element(f"{{{foreign_namespace}}}item"))
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            rewritten.writestr(entry, payload)
    path.write_bytes(output.getvalue())

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("arbitrary content-types child must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_content_types_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_checks_later_nonstandard_worksheet_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "sheet-1"
    for index in range(2, MAX_XLSX_SHEETS + 2):
        workbook.create_sheet(title=f"Sheet{index}")["A1"] = f"sheet-{index}"
    workbook.save(path)
    workbook.close()
    _retarget_later_worksheet_to_entity_payload(
        path,
        sheet_index=MAX_XLSX_SHEETS + 1,
    )
    with zipfile.ZipFile(path, "r") as archive:
        later_path = f"xl/worksheets/sheet{MAX_XLSX_SHEETS + 1}.payload"
        assert later_path in archive.namelist()
        assert b"<!DOCTYPE" in archive.read(later_path)

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("later worksheet declarations must fail before openpyxl size parsing")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_xml_entities_unsupported"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_xlsx_parser_rejects_unreferenced_duplicate_worksheet_target_before_openpyxl(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)
    _inject_unreferenced_duplicate_worksheet_target(path)
    with zipfile.ZipFile(path, "r") as archive:
        relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        worksheet_targets = [
            child.attrib["Target"]
            for child in relationships
            if child.attrib.get("Type", "").endswith("/worksheet")
        ]
        assert len(worksheet_targets) != len(set(worksheet_targets))

    def fail_load_workbook(*_args, **_kwargs):
        raise AssertionError("duplicate worksheet target must fail before openpyxl")

    monkeypatch.setattr("openpyxl.load_workbook", fail_load_workbook)

    with pytest.raises(
        AttachmentPreprocessingError,
        match="xlsx_relationship_structure_unsupported",
    ):
        parse_xlsx_attachment(path=path, requirement=_requirement())


@pytest.mark.parametrize("utf16", [False, True], ids=["utf8", "utf16"])
def test_xlsx_parser_rejects_dtd_and_entity_declarations(tmp_path, utf16):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)
    _inject_worksheet_entity_declaration(path, utf16=utf16)

    with pytest.raises(AttachmentPreprocessingError, match="xlsx_xml_entities_unsupported"):
        parse_xlsx_attachment(path=path, requirement=_requirement())


@pytest.mark.parametrize(
    ("payload", "expected_code"),
    [
        (b"not-a-workbook", "xlsx_parse_failed"),
        (b"x" * (MAX_XLSX_FILE_BYTES + 1), "attachment_parser_file_too_large"),
    ],
    ids=["malformed", "oversized"],
)
def test_xlsx_parser_fails_truthfully_for_malformed_or_oversized_input(tmp_path, payload, expected_code):
    path = tmp_path / "book.xlsx"
    path.write_bytes(payload)

    with pytest.raises(AttachmentPreprocessingError, match=expected_code):
        parse_xlsx_attachment(path=path, requirement=_requirement())


def test_platform_registry_marks_legacy_workbook_unsupported():
    contract = build_attachment_preprocessing_contract(
        file_ids=["file-a"],
        file_names=["legacy.xls"],
    )
    requirement = attachment_requirements_from_contract(contract)[0]

    assert requirement.supported is False
    assert requirement.parser_id == "unsupported"
    assert is_known_binary_workbook(file_name="legacy.xls") is True


def test_platform_registry_uses_server_content_type_when_extension_is_generic(tmp_path):
    path = tmp_path / "attachment.bin"
    _write_workbook(path)
    contract = build_attachment_preprocessing_contract(
        file_ids=["file-a"],
        file_names=["attachment.bin"],
        content_types=["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
    )
    requirement = attachment_requirements_from_contract(contract)[0]

    parsed = parse_xlsx_attachment(path=path, requirement=requirement)

    assert requirement.extension == ".bin"
    assert requirement.content_type.endswith("spreadsheetml.sheet")
    assert parsed.evidence.status == "parsed"


def test_parser_rejects_brokered_bytes_that_do_not_match_worker_materialization(tmp_path):
    path = tmp_path / "book.xlsx"
    materialized = b"AAAA"
    path.write_bytes(materialized)
    contract = build_attachment_preprocessing_contract(
        attachment_facts=[
            MaterializedAttachmentFact(
                file_id="file-a",
                file_name="book.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                byte_count=len(materialized),
                sha256=hashlib.sha256(materialized).hexdigest(),
            )
        ],
    )
    requirement = attachment_requirements_from_contract(contract)[0]
    assert requirement.expected_byte_count == len(materialized)
    assert requirement.expected_sha256 == hashlib.sha256(materialized).hexdigest()

    path.write_bytes(b"BBBB")
    with pytest.raises(AttachmentPreprocessingError, match="attachment_parser_staged_file_mismatch"):
        parse_xlsx_attachment(path=path, requirement=requirement)


def test_duplicate_xlsx_basenames_keep_distinct_file_facts_and_requirements():
    first = b"AAAA"
    second = b"BBBB"
    contract = build_attachment_preprocessing_contract(
        attachment_facts=[
            MaterializedAttachmentFact(
                file_id="file-a",
                file_name="book.xlsx",
                content_type="application/octet-stream",
                byte_count=len(first),
                sha256=hashlib.sha256(first).hexdigest(),
            ),
            MaterializedAttachmentFact(
                file_id="file-b",
                file_name="book.xlsx",
                content_type="application/octet-stream",
                byte_count=len(second),
                sha256=hashlib.sha256(second).hexdigest(),
            ),
        ]
    )

    requirements = attachment_requirements_from_contract(contract)

    assert [requirement.file_id for requirement in requirements] == ["file-a", "file-b"]
    assert [requirement.file_name for requirement in requirements] == ["book.xlsx", "book.xlsx"]
    assert requirements[0].expected_byte_count == requirements[1].expected_byte_count
    assert requirements[0].expected_sha256 != requirements[1].expected_sha256


def test_worker_evidence_validation_rejects_mismatch_and_accepts_exact_record(tmp_path):
    path = tmp_path / "book.xlsx"
    _write_workbook(path)
    requirement = _requirement()
    parsed = parse_xlsx_attachment(path=path, requirement=requirement)
    evidence = parsed.evidence.model_dump(mode="json")

    assert validate_required_parser_evidence(
        requirements=[requirement],
        evidence=[evidence],
    ) == (True, "")

    evidence["parser_version"] = "999"
    assert validate_required_parser_evidence(
        requirements=[requirement],
        evidence=[evidence],
    ) == (False, "attachment_parser_evidence_mismatch")
