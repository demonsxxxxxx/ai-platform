from pathlib import Path
import importlib.util
import json

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]


def load_fill_excel_module():
    script = ROOT / "skills" / "audit-finding-rca" / "scripts" / "fill_excel.py"
    spec = importlib.util.spec_from_file_location("audit_finding_rca_fill_excel", script)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load {script}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def create_workbook(path: Path, rows: list[list[str | None]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_audit_finding_rca_scan_and_fill_creates_output_columns_without_overwriting(tmp_path, monkeypatch):
    module = load_fill_excel_module()
    monkeypatch.chdir(tmp_path)
    source = Path("audit-findings.xlsx")

    create_workbook(
        source,
        [
            ["缺陷描述", "客户法规引用"],
            ["SOP 未明确批记录自动化系统变更后的同步评估要求", "GMP 第八章"],
            ["设备确认报告中未覆盖报警确认", "ICH Q7 §12"],
        ],
    )

    scan = module.scan_empty_rows(source)
    assert scan["empty_count"] == 2
    assert scan["excel_path"] == "audit-findings.xlsx"
    assert scan["rca_created"] is True
    assert scan["capa_created"] is True

    result = module.fill_excel(
        source,
        {
            "2": {"rca": "RCA-2", "capa": "CAPA-2"},
            "3": {"rca": "RCA-3", "capa": "CAPA-3"},
        },
        "output",
    )

    assert Path(result["output_path"]).is_file()
    assert Path(result["output_path"]).parts[0] == "output"
    original = openpyxl.load_workbook(source).active
    assert original.max_column == 2

    filled = openpyxl.load_workbook(result["output_path"]).active
    assert filled.cell(row=1, column=3).value == "原因分析 (RCA)"
    assert filled.cell(row=1, column=4).value == "整改计划 (CAPA)"
    assert filled.cell(row=2, column=3).value == "RCA-2"
    assert filled.cell(row=2, column=4).value == "CAPA-2"


def test_audit_finding_rca_fill_skips_existing_cells(tmp_path, monkeypatch):
    module = load_fill_excel_module()
    monkeypatch.chdir(tmp_path)
    source = Path("audit-findings-existing.xlsx")

    create_workbook(
        source,
        [
            ["缺陷描述", "原因分析", "整改计划"],
            ["SOP 记录字段有待补充", "已有RCA", None],
        ],
    )

    result = module.fill_excel(
        source,
        {"2": {"rca": "新RCA", "capa": "新CAPA"}},
        "output",
    )

    filled = openpyxl.load_workbook(result["output_path"]).active
    assert filled.cell(row=2, column=2).value == "已有RCA"
    assert filled.cell(row=2, column=3).value == "新CAPA"
    assert result["rca_filled"] == 0
    assert result["capa_filled"] == 1


def test_audit_finding_rca_fill_escapes_formula_like_text(tmp_path, monkeypatch):
    module = load_fill_excel_module()
    monkeypatch.chdir(tmp_path)
    source = Path("audit-findings-formula.xlsx")
    create_workbook(source, [["缺陷描述"], ["公式注入风险"]])

    result = module.fill_excel(
        source,
        {
            "2": {
                "rca": '=HYPERLINK("https://example.invalid","open")',
                "capa": "+SUM(1,1)",
            }
        },
        "output",
    )

    filled = openpyxl.load_workbook(result["output_path"], data_only=False).active
    assert filled.cell(row=2, column=2).value == "'=HYPERLINK(\"https://example.invalid\",\"open\")"
    assert filled.cell(row=2, column=2).data_type == "s"
    assert filled.cell(row=2, column=3).value == "'+SUM(1,1)"
    assert filled.cell(row=2, column=3).data_type == "s"


def test_audit_finding_rca_rejects_paths_outside_workspace_and_output_dir(tmp_path, monkeypatch):
    module = load_fill_excel_module()
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    monkeypatch.chdir(workspace)

    source = Path("audit-findings.xlsx")
    create_workbook(source, [["缺陷描述"], ["边界检查"]])
    outside_workbook = tmp_path / "outside.xlsx"
    create_workbook(outside_workbook, [["缺陷描述"], ["不应读取"]])
    outside_data = tmp_path / "rca_data.json"
    outside_data.write_text(json.dumps({"2": {"rca": "RCA", "capa": "CAPA"}}), encoding="utf-8")

    with pytest.raises(ValueError, match="must stay inside the workspace"):
        module.scan_empty_rows(outside_workbook)
    with pytest.raises(ValueError, match="must stay inside the workspace"):
        module.load_rca_data(outside_data)
    with pytest.raises(ValueError, match="must be under output"):
        module.fill_excel(source, {"2": {"rca": "RCA", "capa": "CAPA"}}, "exports")
    with pytest.raises(ValueError, match="must be under output"):
        module._write_json("scan.json", {"ok": True})


def test_audit_finding_rca_cli_reports_relative_output_json_path(tmp_path, monkeypatch, capsys):
    module = load_fill_excel_module()
    monkeypatch.chdir(tmp_path)
    source = Path("audit-findings.xlsx")
    create_workbook(source, [["缺陷描述"], ["输出路径不应暴露绝对目录"]])
    output_json = (tmp_path / "output" / "scan.json").resolve()

    monkeypatch.setattr(
        "sys.argv",
        [
            "fill_excel.py",
            "--scan",
            "--excel",
            str(source),
            "--output-json",
            str(output_json),
        ],
    )

    assert module.main() == 0

    captured = capsys.readouterr()
    assert "[OK] JSON written: output" in captured.out
    assert str(tmp_path) not in captured.out
