#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Scan and fill audit finding Excel workbooks with RCA/CAPA text."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import re
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RCA_HEADER = "原因分析 (RCA)"
CAPA_HEADER = "整改计划 (CAPA)"
FORMULA_PREFIXES = ("=", "+", "-", "@")


def _workspace_root() -> Path:
    return Path.cwd().resolve()


def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
    except ValueError:
        return False
    return True


def _display_path(path: Path) -> str:
    try:
        return str(path.resolve(strict=False).relative_to(_workspace_root()))
    except ValueError:
        return path.name


def _resolve_workspace_path(path: str | Path, label: str) -> Path:
    root = _workspace_root()
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = root / candidate
    resolved = candidate.resolve(strict=False)
    if not _is_relative_to(resolved, root):
        raise ValueError(f"{label} must stay inside the workspace")
    return resolved


def _resolve_output_path(path: str | Path, label: str) -> Path:
    resolved = _resolve_workspace_path(path, label)
    output_root = (_workspace_root() / "output").resolve(strict=False)
    if resolved != output_root and not _is_relative_to(resolved, output_root):
        raise ValueError(f"{label} must be under output")
    return resolved


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _excel_text(value: Any) -> str:
    text = _cell_text(value)
    if text.startswith(FORMULA_PREFIXES):
        return f"'{text}"
    return text


def _header_text(ws, column: int) -> str:
    value = ws.cell(row=1, column=column).value
    return _cell_text(value) or f"Column {get_column_letter(column)}"


def _matches_any(text: str, patterns: list[str]) -> bool:
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in patterns)


def detect_columns(ws) -> dict[str, Any]:
    """Detect RCA/CAPA columns from row 1; create output columns if absent."""
    rca_patterns = [
        r"原因分析",
        r"根因",
        r"\bRCA\b",
        r"Root\s*Cause",
        r"Cause\s*Analysis",
    ]
    capa_patterns = [
        r"整改计划",
        r"纠正",
        r"预防措施",
        r"\bCAPA\b",
        r"Corrective",
        r"Preventive",
    ]

    rca_col = None
    capa_col = None
    rca_label = ""
    capa_label = ""

    for column in range(1, ws.max_column + 1):
        text = _cell_text(ws.cell(row=1, column=column).value)
        if not text:
            continue
        if rca_col is None and _matches_any(text, rca_patterns):
            rca_col = column
            rca_label = text
            continue
        if capa_col is None and _matches_any(text, capa_patterns):
            capa_col = column
            capa_label = text

    next_col = ws.max_column + 1
    rca_created = False
    capa_created = False
    if rca_col is None:
        rca_col = next_col
        next_col += 1
        rca_label = RCA_HEADER
        rca_created = True
    if capa_col is None:
        capa_col = next_col
        capa_label = CAPA_HEADER
        capa_created = True

    finding_columns = [
        column
        for column in range(1, ws.max_column + 1)
        if column not in {rca_col, capa_col}
    ]
    if not finding_columns and ws.max_column:
        finding_columns = [1]

    return {
        "rca_col": rca_col,
        "capa_col": capa_col,
        "rca_label": rca_label,
        "capa_label": capa_label,
        "rca_created": rca_created,
        "capa_created": capa_created,
        "finding_columns": finding_columns,
    }


def find_latest_excel(search_root: str | Path = ".") -> Path:
    """Return the newest non-output workbook under the search root."""

    root = _resolve_workspace_path(search_root, "search_root")
    candidates = [
        path
        for path in root.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and "已填写RCA和CAPA" not in path.stem
    ]
    if not candidates:
        raise FileNotFoundError(f"No input .xlsx file found under {_display_path(root)}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _row_finding_text(ws, row: int, columns: list[int]) -> tuple[str, dict[str, str]]:
    source_values: dict[str, str] = {}
    parts: list[str] = []
    for column in columns:
        value = _cell_text(ws.cell(row=row, column=column).value)
        if not value:
            continue
        label = _header_text(ws, column)
        source_values[label] = value
        parts.append(f"{label}: {value}")
    return "\n".join(parts), source_values


def scan_empty_rows(excel_path: str | Path) -> dict[str, Any]:
    """Describe workbook rows that still need RCA or CAPA content."""

    workbook_path = _resolve_workspace_path(excel_path, "excel_path")
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    cols = detect_columns(ws)
    rca_col = cols["rca_col"]
    capa_col = cols["capa_col"]

    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        finding_text, source_values = _row_finding_text(ws, row, cols["finding_columns"])
        if not finding_text:
            continue
        rca_empty = cols["rca_created"] or _is_blank(ws.cell(row=row, column=rca_col).value)
        capa_empty = cols["capa_created"] or _is_blank(ws.cell(row=row, column=capa_col).value)
        if rca_empty or capa_empty:
            rows.append(
                {
                    "row": row,
                    "finding_text": finding_text,
                    "source_values": source_values,
                    "rca_empty": rca_empty,
                    "capa_empty": capa_empty,
                }
            )

    return {
        "excel_path": _display_path(workbook_path),
        "sheet": ws.title,
        "total_rows": ws.max_row,
        "empty_count": len(rows),
        "rows": rows,
        "rca_col": rca_col,
        "capa_col": capa_col,
        "rca_label": cols["rca_label"],
        "capa_label": cols["capa_label"],
        "rca_created": cols["rca_created"],
        "capa_created": cols["capa_created"],
    }


def load_rca_data(data_path: str | Path) -> dict[str, dict[str, str]]:
    """Load row-keyed RCA/CAPA text from a JSON file."""

    path = _resolve_workspace_path(data_path, "data_path")
    with path.open("r", encoding="utf-8-sig") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("RCA data JSON must be an object keyed by Excel row number")
    normalized: dict[str, dict[str, str]] = {}
    for row, content in payload.items():
        if not isinstance(content, dict):
            raise ValueError(f"RCA data row {row!r} must be an object")
        normalized[str(int(row))] = {
            "rca": _cell_text(content.get("rca")),
            "capa": _cell_text(content.get("capa")),
        }
    return normalized


def _unique_output_path(output_dir: Path, source_path: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    candidate = output_dir / f"{source_path.stem}_已填写RCA和CAPA{source_path.suffix}"
    if not candidate.exists():
        return candidate
    for index in range(2, 1000):
        candidate = output_dir / f"{source_path.stem}_已填写RCA和CAPA-{index}{source_path.suffix}"
        if not candidate.exists():
            return candidate
    raise FileExistsError("Too many output workbook versions already exist")


def _apply_column_formatting(ws, rca_col: int, capa_col: int) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    for column in [rca_col, capa_col]:
        ws.column_dimensions[get_column_letter(column)].width = 85
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=column).alignment = wrap

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fill_excel(
    excel_path: str | Path,
    rca_data: dict[str, dict[str, str]],
    output_dir: str | Path = "output",
) -> dict[str, Any]:
    """Write RCA/CAPA text to a copied workbook without overwriting source cells."""

    source_path = _resolve_workspace_path(excel_path, "excel_path")
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active
    cols = detect_columns(ws)
    rca_col = cols["rca_col"]
    capa_col = cols["capa_col"]

    if cols["rca_created"]:
        ws.cell(row=1, column=rca_col).value = RCA_HEADER
    if cols["capa_created"]:
        ws.cell(row=1, column=capa_col).value = CAPA_HEADER

    rca_filled = 0
    capa_filled = 0
    skipped_existing = 0
    skipped_out_of_range = 0

    for row_text, content in sorted(rca_data.items(), key=lambda item: int(item[0])):
        row = int(row_text)
        if row < 2 or row > ws.max_row:
            skipped_out_of_range += 1
            continue
        rca_cell = ws.cell(row=row, column=rca_col)
        capa_cell = ws.cell(row=row, column=capa_col)

        if _is_blank(rca_cell.value):
            rca_cell.value = _excel_text(content.get("rca", ""))
            rca_filled += 1
        else:
            skipped_existing += 1
        if _is_blank(capa_cell.value):
            capa_cell.value = _excel_text(content.get("capa", ""))
            capa_filled += 1
        else:
            skipped_existing += 1

    _apply_column_formatting(ws, rca_col, capa_col)
    output_path = _unique_output_path(_resolve_output_path(output_dir, "output_dir"), source_path)
    wb.save(output_path)

    return {
        "input_path": _display_path(source_path),
        "output_path": _display_path(output_path),
        "sheet": ws.title,
        "total_rows": ws.max_row,
        "rca_col": rca_col,
        "capa_col": capa_col,
        "rca_label": cols["rca_label"],
        "capa_label": cols["capa_label"],
        "rca_created": cols["rca_created"],
        "capa_created": cols["capa_created"],
        "rca_filled": rca_filled,
        "capa_filled": capa_filled,
        "skipped_existing_cells": skipped_existing,
        "skipped_out_of_range_rows": skipped_out_of_range,
    }


def _write_json(path: str | Path, payload: dict[str, Any]) -> None:
    target = _resolve_output_path(path, "output_json")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    """Run the command-line scanner or filler."""

    parser = argparse.ArgumentParser(description="Scan or fill audit finding RCA/CAPA Excel workbooks.")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--scan", action="store_true", help="Scan rows needing RCA/CAPA.")
    mode.add_argument("--fill", action="store_true", help="Fill RCA/CAPA from JSON data.")
    parser.add_argument("--excel", help="Input .xlsx workbook. Defaults to latest workbook in --search-root.")
    parser.add_argument("--search-root", default=".", help="Directory used when --excel is omitted.")
    parser.add_argument("--output-json", help="Write scan/fill result JSON to this path.")
    parser.add_argument("--data", default="output/rca_data.json", help="RCA/CAPA JSON for --fill.")
    parser.add_argument("--output-dir", default="output", help="Directory for filled workbook output.")
    args = parser.parse_args()

    excel_path = Path(args.excel) if args.excel else find_latest_excel(args.search_root)
    if args.fill:
        result = fill_excel(excel_path, load_rca_data(args.data), args.output_dir)
    else:
        result = scan_empty_rows(excel_path)

    if args.output_json:
        _write_json(args.output_json, result)
        print(f"[OK] JSON written: {args.output_json}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
