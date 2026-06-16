import openpyxl
p = r'C:\Users\Yanan.Liu\Desktop\3.2.S.7.3 申报资料智能体测试范例-真实数据\3.2.S.7.3-IP350\参考文件4-稳定性试验数据汇总.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
print(wb.sheetnames)
for s in ['原液长期', '原液加速']:
    ws = wb[s]
    print('---', s, ws.max_row, ws.max_column)
    for r in range(1, min(ws.max_row, 260) + 1):
        vals = [ws.cell(r, c).value for c in range(1, 18)]
        if any(v not in (None, '') for v in vals):
            print(r, vals)
